#!/usr/bin/env python3
# -*- coding: utf-8 -*-


import argparse
import glob
import json
from openpyxl import Workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

parser = argparse.ArgumentParser(description="Script options")
parser.add_argument("--input", help="An input option")
parser.add_argument("--output")
args = parser.parse_args()

def flatten_extend(matrix):
    flat_list = []
    for row in matrix:
        if type(row) is list:
            flat_list.extend(row)
        else:
            flat_list.append(row)
    return flat_list


def main(output):

    wb = Workbook()
    ft = Font(name="Sans", bold=True)
    cb_even = PatternFill(fill_type="solid", fgColor="d9e1f2")
    cb_uneven = PatternFill(fill_type="solid", fgColor="cdd1d9")
    cb_failed = PatternFill(fill_type="solid", fgColor="FF3300")

    toolchains = []
    bucket = {}

    reports = sorted(glob.glob("*.json"))

    # All the JSON reports
    for report in reports:

        # Parse the JSON file
        with open(report) as f:
            data = json.load(f)

        sample = data["sample"]
        matches = data["matches"]

        # group by rule, then sample
        for match in matches:
            toolchain = match["toolchain"]
            rule = match["rule"]

            if not toolchain in toolchains:
                toolchains.append(toolchain)

            if rule not in bucket:
                bucket[rule] = {}

            if sample in bucket[rule]:
                bucket[rule][sample].append(match)
            else:
                bucket[rule][sample] = [ match ]

    toolchains.sort()

    sheet_index = 0

    # write one sheet per rule, listing all the samples
    for rule, samples in bucket.items():
         
        ws = wb.create_sheet(title=rule, index=sheet_index)
        sheet_index += 1

        header = [ "" ] 
        for tool in toolchains:
            header.append( [tool, "", ""])
        ws.append(flatten_extend(header))

        header = [ "Probe" ]
        for tool in toolchains:
            header.append([ "% GMO","Reads WT","Reads GMO"])

        ws.append(flatten_extend(header))

        row = 0

        for sample, reports in samples.items():

            row += 1

            this_row = [ sample ]

            failed = False

            for tool in toolchains:

                perc_gmo = ""
                ref_cov = ""
                alt_cov = ""

                report = [d for d in reports if d['toolchain'] == tool ][0]

                if report:
                    perc_gmo = float(report["perc_gmo"])
                    ref_cov = int(report["ref_cov"])
                    alt_cov = int(report["alt_cov"])
                    if ref_cov == "NA" and "bam_cov" in report:
                        ref_cov = report["bam_cov"]
                        alt_cov = "-"
                
                # Simplistic rule to catch failed samples. 
                if ref_cov < 100:
                    failed = True

                for e in [ perc_gmo, ref_cov, alt_cov ]:
                    this_row.append(e)

            ws.append(this_row)

            if failed:
                bg_color = cb_failed
            else:
                bg_color = cb_even if (row & 1) else cb_uneven
                
            for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
                ws[col+str(ws._current_row)].fill = bg_color       

         # Auto-width for columns
        dim_holder = DimensionHolder(worksheet=ws)
        for column in range(ws.min_column, ws.max_column + 1):
            dim_holder[get_column_letter(column)] = ColumnDimension(ws, min=column, max=column, width=20)
        ws.column_dimensions = dim_holder
        ws.freeze_panes = ws["A3"]

    # Write excel file
    wb.save(output)
    

if __name__ == '__main__':
    main(args.output)
